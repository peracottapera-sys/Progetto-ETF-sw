"""
import_etf_railway.py
Importa il catalogo ETF da JustETF (Excel) nel DB PostgreSQL di Railway.

Uso:
  pip install psycopg2-binary openpyxl
  python import_etf_railway.py
"""

import psycopg2
import openpyxl
import os
import sys

# ── CONFIGURA QUESTO PERCORSO SE NECESSARIO ──────────────────────────────────
EXCEL_PATH = r"D:\Documenti\Casa\ETF_App\Lista_ETF_-_032026.xlsx"
DB_URL     = "postgresql://postgres:JZKhCmNKgtZZfdDfQSPmgwORwBgxuAHO@crossover.proxy.rlwy.net:20706/railway"
# ─────────────────────────────────────────────────────────────────────────────


def safe_float(val, multiplier=1.0):
    if val is None or val == '-' or val == '' or val == 'Graph':
        return None
    try:
        return round(float(val) * multiplier, 4)
    except (ValueError, TypeError):
        return None


def is_valid_isin(isin):
    if not isin or not isinstance(isin, str):
        return False
    isin = isin.strip()
    return len(isin) == 12 and isin[:2].isalpha() and isin[2:].isalnum()


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERRORE: File Excel non trovato: {EXCEL_PATH}")
        sys.exit(1)

    print(f"Lettura Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    total_rows = ws.max_row - 1
    print(f"Righe trovate: {total_rows}")

    print(f"Connessione a Railway PostgreSQL...")
    try:
        conn = psycopg2.connect(DB_URL, sslmode='require', connect_timeout=15)
        cur = conn.cursor()
        print("Connesso!")
    except Exception as e:
        print(f"ERRORE connessione: {e}")
        sys.exit(1)

    # Crea tabella se non esiste (compatibile con server.js)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS etf_catalog (
            isin            TEXT PRIMARY KEY,
            name            TEXT,
            emittente       TEXT,
            ter             REAL,
            valuta          TEXT,
            aum_mln         REAL,
            perf1m          REAL,
            perf6m          REAL,
            perf1y          REAL,
            perf3y          REAL,
            perf5y          REAL,
            vol1y           REAL,
            maxdd1y         REAL,
            maxdd5y         REAL,
            distribuzione   TEXT,
            replica         TEXT,
            ticker_yahoo    TEXT,
            categoria       TEXT,
            quotazione      REAL,
            active          INTEGER DEFAULT 1,
            updated_at      TEXT
        )
    """)
    cur.execute("ALTER TABLE etf_catalog ADD COLUMN IF NOT EXISTS maxdd5y REAL")
    conn.commit()
    print("Tabella etf_catalog pronta.")

    cur.execute("SELECT COUNT(*) FROM etf_catalog")
    count_before = cur.fetchone()[0]
    print(f"ETF gia presenti nel DB: {count_before}")

    inserted = updated = skipped = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Stessa struttura colonne dello script SQLite originale
        name, valuta, aum, ter, _, p1m, p3m, p6m, p1y, p3y, p5y, ytd, \
        v1y, v3y, v5y, dd1y, dd3y, dd5y, distr, replica, isin = row

        if not is_valid_isin(isin):
            skipped += 1
            continue
        if not name or str(name).startswith('Annuncio'):
            skipped += 1
            continue

        isin = isin.strip()
        name = str(name).strip()

        try:
            cur.execute("""
                INSERT INTO etf_catalog
                    (isin, name, valuta, aum_mln, ter,
                     perf1m, perf6m, perf1y, perf3y, perf5y,
                     vol1y, maxdd1y, maxdd5y,
                     distribuzione, replica, active, updated_at)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,1,'2026-03-18')
                ON CONFLICT (isin) DO UPDATE SET
                    name          = EXCLUDED.name,
                    valuta        = EXCLUDED.valuta,
                    aum_mln       = EXCLUDED.aum_mln,
                    ter           = EXCLUDED.ter,
                    perf1m        = EXCLUDED.perf1m,
                    perf6m        = EXCLUDED.perf6m,
                    perf1y        = EXCLUDED.perf1y,
                    perf3y        = EXCLUDED.perf3y,
                    perf5y        = EXCLUDED.perf5y,
                    vol1y         = EXCLUDED.vol1y,
                    maxdd1y       = EXCLUDED.maxdd1y,
                    maxdd5y       = EXCLUDED.maxdd5y,
                    distribuzione = EXCLUDED.distribuzione,
                    replica       = EXCLUDED.replica,
                    updated_at    = '2026-03-18'
            """, (
                isin, name,
                str(valuta).strip() if valuta else None,
                safe_float(aum),
                safe_float(ter, 100),     # 0.0007 -> 0.07%
                safe_float(p1m, 100),
                safe_float(p6m, 100),
                safe_float(p1y, 100),
                safe_float(p3y, 100),
                safe_float(p5y, 100),
                safe_float(v1y, 100),
                safe_float(dd1y, 100),
                safe_float(dd5y, 100),
                str(distr).strip()   if distr   else None,
                str(replica).strip() if replica else None,
            ))

            if cur.rowcount == 1:
                inserted += 1
            else:
                updated += 1

        except Exception as e:
            print(f"  Errore riga {isin}: {e}")
            skipped += 1

        if (i - 1) % 500 == 0 and (i - 1) > 0:
            print(f"  Elaborati {i-1}/{total_rows}...")
            conn.commit()

    conn.commit()

    cur.execute("SELECT COUNT(*) FROM etf_catalog")
    count_after = cur.fetchone()[0]

    cur.close()
    conn.close()

    print(f"\nImport completato!")
    print(f"   ETF nel DB prima:  {count_before}")
    print(f"   ETF nel DB dopo:   {count_after}")
    print(f"   Inseriti:          {inserted}")
    print(f"   Aggiornati:        {updated}")
    print(f"   Saltati:           {skipped}")
    print(f"\nRailway ha ora il catalogo ETF. Riprova l'algoritmo AI!")


if __name__ == "__main__":
    main()
