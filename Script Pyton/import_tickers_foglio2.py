"""
import_tickers_foglio2.py
Importa ticker Yahoo dal Foglio2 dell'Excel JustETF nel DB Railway.
Usa il Foglio1 come base (ISIN di riferimento) e aggiunge ticker dal Foglio2.

Uso:
  pip install psycopg2-binary openpyxl requests
  python import_tickers_foglio2.py
"""

import psycopg2
import openpyxl
import requests
import time
import os
import sys

DB_URL    = "postgresql://postgres:JZKhCmNKgtZZfdDfQSPmgwORwBgxuAHO@crossover.proxy.rlwy.net:20706/railway"
EXCEL_PATH = r"D:\Documenti\Casa\ETF_App\Lista_ETF_-_032026.xlsx"

HEADERS = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}
SUFFISSI_BORSA = ['.MI', '.AS', '.DE', '.PA', '.L', '.F', '.SW', '.BR', '.AX']

def test_ticker(ticker):
    """Verifica che il ticker esista su Yahoo e restituisce il prezzo."""
    try:
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?interval=1d&range=5d"
        r = requests.get(url, headers=HEADERS, timeout=8)
        data = r.json()
        prezzo = data.get('chart', {}).get('result', [{}])[0].get('meta', {}).get('regularMarketPrice')
        if prezzo and prezzo > 0:
            return prezzo
    except:
        pass
    return None

def trova_ticker_yahoo(ticker_base):
    """Prova il ticker con vari suffissi di borsa e restituisce (ticker_completo, prezzo)."""
    if not ticker_base:
        return None, None
    # Prova prima senza suffisso (alcuni ticker sono già completi)
    if '.' in ticker_base:
        prezzo = test_ticker(ticker_base)
        if prezzo:
            return ticker_base, prezzo
    # Prova con suffissi
    for suf in SUFFISSI_BORSA:
        t = ticker_base + suf
        prezzo = test_ticker(t)
        if prezzo:
            return t, prezzo
        time.sleep(0.1)
    return None, None

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERRORE: File non trovato: {EXCEL_PATH}")
        sys.exit(1)

    print("Lettura Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Leggi Foglio1 — ISIN di riferimento (base)
    ws1 = wb['Foglio1']
    isin_base = set()
    for row in ws1.iter_rows(min_row=2, values_only=True):
        isin = str(row[20]).strip() if row[20] else ''
        if len(isin) == 12 and isin[:2].isalpha():
            isin_base.add(isin)
    print(f"ISIN nel Foglio1 (base): {len(isin_base)}")

    # Leggi Foglio2 — ticker + domicilio + maxDD max
    # Header: ['', 'Nome', 'Grafico', 'Max drawdown MAX', 'Domicilio', 'ISIN', 'Ticker', 'WKN', 'Valor']
    ws2 = wb['Foglio2']
    ticker_map = {}  # isin -> {ticker, domicilio, maxdd_max}
    for row in ws2.iter_rows(min_row=2, values_only=True):
        isin    = str(row[5]).strip() if row[5] else ''
        ticker  = str(row[6]).strip() if row[6] else ''
        domicilio = str(row[4]).strip() if row[4] else ''
        maxdd_max = row[3]
        if len(isin) == 12 and isin[:2].isalpha() and ticker and ticker != 'None':
            if isin not in ticker_map:  # prendi solo la prima occorrenza
                ticker_map[isin] = {
                    'ticker': ticker,
                    'domicilio': domicilio,
                    'maxdd_max': float(maxdd_max) * 100 if maxdd_max and str(maxdd_max) not in ('None', '', '-') else None
                }
    print(f"ISIN con ticker nel Foglio2: {len(ticker_map)}")

    # Solo ISIN presenti in entrambi i fogli
    isin_da_aggiornare = {k: v for k, v in ticker_map.items() if k in isin_base}
    print(f"ISIN da aggiornare (in entrambi i fogli): {len(isin_da_aggiornare)}")

    # Connessione DB
    print("\nConnessione a Railway PostgreSQL...")
    conn = psycopg2.connect(DB_URL, sslmode='require', connect_timeout=15)
    cur = conn.cursor()
    print("Connesso!\n")

    # Controlla quali ISIN hanno già ticker nel DB
    cur.execute("SELECT isin FROM etf_catalog WHERE ticker_yahoo IS NOT NULL AND ticker_yahoo != ''")
    gia_con_ticker = set(r[0] for r in cur.fetchall())
    print(f"ETF già con ticker nel DB: {len(gia_con_ticker)}")

    da_verificare = {k: v for k, v in isin_da_aggiornare.items() if k not in gia_con_ticker}
    print(f"ETF da verificare su Yahoo: {len(da_verificare)}")
    print(f"Stima tempo: ~{len(da_verificare) * 0.5 / 60:.0f} minuti\n")
    print("Premi CTRL+C per interrompere — i progressi vengono salvati.\n")

    trovati = 0
    non_trovati = 0
    oggi = __import__('datetime').date.today().isoformat()

    for i, (isin, info) in enumerate(da_verificare.items(), 1):
        ticker_base = info['ticker']
        ticker_yahoo, prezzo = trova_ticker_yahoo(ticker_base)

        if ticker_yahoo and prezzo:
            cur.execute("""
                UPDATE etf_catalog 
                SET ticker_yahoo = %s, quotazione = %s
                WHERE isin = %s
            """, [ticker_yahoo, prezzo, isin])
            # Salva prezzo storico
            cur.execute("""
                INSERT INTO prezzi_storici (isin, data, prezzo) VALUES (%s, %s, %s)
                ON CONFLICT (isin, data) DO UPDATE SET prezzo = EXCLUDED.prezzo
            """, [isin, oggi, prezzo])
            conn.commit()
            trovati += 1
            print(f"[{i}/{len(da_verificare)}] ✓ {isin} → {ticker_yahoo} @ €{prezzo:.2f}")
        else:
            non_trovati += 1
            if i % 100 == 0:
                print(f"[{i}/{len(da_verificare)}] ... {trovati} trovati, {non_trovati} non trovati")

        time.sleep(0.3)

    # Aggiorna anche domicilio e maxdd_max per tutti (anche quelli già con ticker)
    print("\nAggiornamento domicilio e maxDrawdown MAX per tutti gli ETF...")
    aggiornati_meta = 0
    for isin, info in isin_da_aggiornare.items():
        if info['maxdd_max'] is not None:
            cur.execute("""
                UPDATE etf_catalog SET maxdd5y = %s WHERE isin = %s AND maxdd5y IS NULL
            """, [info['maxdd_max'], isin])
            if cur.rowcount > 0:
                aggiornati_meta += 1
    conn.commit()
    print(f"MaxDD MAX aggiornato per {aggiornati_meta} ETF")

    cur.close()
    conn.close()

    print(f"\n✅ Completato!")
    print(f"   Ticker trovati e salvati: {trovati}")
    print(f"   ETF senza ticker Yahoo:   {non_trovati}")
    print(f"\nOra l'algoritmo AI avrà molti più ETF disponibili!")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nInterrotto. I progressi sono stati salvati.")
        sys.exit(0)
EOF