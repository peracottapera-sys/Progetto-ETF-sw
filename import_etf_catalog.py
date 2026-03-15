"""
import_etf_catalog.py
Importa il catalogo ETF da JustETF (Excel) nel DB SQLite dell'app.

Uso:
  python import_etf_catalog.py

Percorsi da configurare:
  EXCEL_PATH  → file Excel scaricato da JustETF
  DB_PATH     → database SQLite del server Node
"""

import sqlite3
import openpyxl
import os
import sys

# ── CONFIGURA QUESTI PERCORSI ────────────────────────────────────────────────
EXCEL_PATH = r"D:\Documenti\Casa\ETF_App\Lista_ETF_-_032026.xlsx"
DB_PATH    = r"D:\Documenti\Casa\ETF_App\etf-server\etf_app.db"
# ─────────────────────────────────────────────────────────────────────────────


def safe_float(val, multiplier=1.0):
    """Converte in float, restituisce None se il valore non è numerico."""
    if val is None or val == '-' or val == '' or val == 'Graph':
        return None
    try:
        return round(float(val) * multiplier, 4)
    except (ValueError, TypeError):
        return None


def is_valid_isin(isin):
    """Verifica formato ISIN base: 2 lettere + 10 caratteri alfanumerici."""
    if not isin or not isinstance(isin, str):
        return False
    isin = isin.strip()
    return len(isin) == 12 and isin[:2].isalpha() and isin[2:].isalnum()


def main():
    # Verifica file
    if not os.path.exists(EXCEL_PATH):
        print(f"ERRORE: File Excel non trovato: {EXCEL_PATH}")
        sys.exit(1)
    if not os.path.exists(DB_PATH):
        print(f"ERRORE: DB SQLite non trovato: {DB_PATH}")
        sys.exit(1)

    print(f"Lettura Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    total_rows = ws.max_row - 1  # escludi header
    print(f"Righe trovate: {total_rows}")

    # Connessione DB
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # Crea tabella etf_catalog se non esiste
    cur.execute("""
        CREATE TABLE IF NOT EXISTS etf_catalog (
            isin            TEXT PRIMARY KEY,
            name            TEXT NOT NULL,
            valuta          TEXT,
            aum_mln         REAL,
            ter             REAL,
            perf1m          REAL,
            perf3m          REAL,
            perf6m          REAL,
            perf1y          REAL,
            perf3y          REAL,
            perf5y          REAL,
            ytd             REAL,
            vol1y           REAL,
            vol3y           REAL,
            vol5y           REAL,
            maxdd1y         REAL,
            maxdd3y         REAL,
            maxdd5y         REAL,
            distribuzione   TEXT,
            replica         TEXT,
            ticker_yahoo    TEXT,
            categoria       TEXT,
            emittente       TEXT,
            active          INTEGER DEFAULT 1,
            updated_at      TEXT DEFAULT (datetime('now'))
        )
    """)

    # Indice per ricerca testuale
    cur.execute("""
        CREATE INDEX IF NOT EXISTS idx_etf_catalog_name
        ON etf_catalog(name)
    """)

    conn.commit()
    print("Tabella etf_catalog pronta.")

    # Import righe
    inserted = 0
    updated  = 0
    skipped  = 0

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        name, valuta, aum, ter, _, p1m, p3m, p6m, p1y, p3y, p5y, ytd, \
        v1y, v3y, v5y, dd1y, dd3y, dd5y, distr, replica, isin = row

        # Filtra righe non valide
        if not is_valid_isin(isin):
            skipped += 1
            continue
        if not name or str(name).startswith('Annuncio'):
            skipped += 1
            continue

        isin  = isin.strip()
        name  = str(name).strip()

        # Le performance nel file sono in decimale (es. 0.1554 = 15.54%)
        # Convertiamo in percentuale
        cur.execute("""
            INSERT INTO etf_catalog
                (isin, name, valuta, aum_mln, ter,
                 perf1m, perf3m, perf6m, perf1y, perf3y, perf5y, ytd,
                 vol1y, vol3y, vol5y, maxdd1y, maxdd3y, maxdd5y,
                 distribuzione, replica)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            ON CONFLICT(isin) DO UPDATE SET
                name          = excluded.name,
                valuta        = excluded.valuta,
                aum_mln       = excluded.aum_mln,
                ter           = excluded.ter,
                perf1m        = excluded.perf1m,
                perf3m        = excluded.perf3m,
                perf6m        = excluded.perf6m,
                perf1y        = excluded.perf1y,
                perf3y        = excluded.perf3y,
                perf5y        = excluded.perf5y,
                ytd           = excluded.ytd,
                vol1y         = excluded.vol1y,
                vol3y         = excluded.vol3y,
                vol5y         = excluded.vol5y,
                maxdd1y       = excluded.maxdd1y,
                maxdd3y       = excluded.maxdd3y,
                maxdd5y       = excluded.maxdd5y,
                distribuzione = excluded.distribuzione,
                replica       = excluded.replica,
                updated_at    = datetime('now')
        """, (
            isin, name,
            str(valuta).strip() if valuta else None,
            safe_float(aum),
            safe_float(ter, 100),        # 0.0007 → 0.07%
            safe_float(p1m,  100),       # 0.0023 → 0.23%
            safe_float(p3m,  100),
            safe_float(p6m,  100),
            safe_float(p1y,  100),
            safe_float(p3y,  100),
            safe_float(p5y,  100),
            safe_float(ytd,  100),
            safe_float(v1y,  100),
            safe_float(v3y,  100),
            safe_float(v5y,  100),
            safe_float(dd1y, 100),
            safe_float(dd3y, 100),
            safe_float(dd5y, 100),
            str(distr).strip()   if distr   else None,
            str(replica).strip() if replica else None,
        ))

        if cur.rowcount == 1:
            inserted += 1
        else:
            updated += 1

        # Progress ogni 500 righe
        if (i - 1) % 500 == 0:
            print(f"  Elaborati {i-1}/{total_rows}...")
            conn.commit()

    conn.commit()
    conn.close()

    print(f"\n✅ Import completato!")
    print(f"   Inseriti: {inserted}")
    print(f"   Aggiornati: {updated}")
    print(f"   Saltati (ISIN invalidi/annunci): {skipped}")
    print(f"\nOra riavvia il server Node per attivare gli endpoint di ricerca.")


if __name__ == "__main__":
    main()
